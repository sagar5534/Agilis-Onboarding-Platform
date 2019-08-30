from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from datetime import date
from forms.models import *
from django.shortcuts import (
    get_object_or_404,
    render,
    redirect,
    HttpResponseRedirect,
)
import random
from django.http import Http404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
import json
from django.core.files.storage import FileSystemStorage
import os
from django.views.decorators.http import require_POST
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.conf import settings
import openpyxl
from openpyxl.drawing.image import Image
import csv
from .models import *
from shutil import copyfile

# Create your views here.

@login_required
def CreateExtensionsForm(request):

    if request.session.get("company"):
        company = request.session.get("company")
        print(company)
    else:
        raise Http404

    try:
        Exts = Extention.objects.filter(company_id=company)
    except Extention.DoesNotExist:
        raise Http404

    tempComp = Company.objects.get(id=company)
    save_path = os.path.join(settings.MEDIA_ROOT, "Reports", str(tempComp.order) + "_Extensions.csv")
    toForm = os.path.join("Reports", str(tempComp.order) + "_Extensions.csv")

    with open(save_path, 'w', newline='') as csvfile:

        fieldnames = ['Extension Number','Extension Name','CIDName','CIDNum','E911CIDNum','DirectorySearchable','DefaultNPA','Wrap Time','Record Call','FwdTime','Record Notify Email','Roaming Passcode','VM Extension Number','VM Forward Email','VM to Email Only','VM Passcode','SIP Username','SIP Password','Extension Active']
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)

        writer.writeheader()

        for Ext in Exts:
            number = Ext.ext
            name = Ext.name
            cidName = Ext.caller_id_name

            try:
                cidNumber = Numbers.objects.get(id=Ext.caller_id_number).number
            except Numbers.DoesNotExist:
                cidNumber = Ext.caller_id_number

            E911Number = cidNumber
            directorySearch = 1
            defaultNPA = "" 
            wrapTime = 0
            recordCall = 2
            forwardTime = 25
            recordNotify = ""
            roamingPass = ""
            VMExtNumber = 0
            VMPasscode = ""
            if (Ext.voicemail == True):
                VMExtNumber = str(1) + str(number)
                VMPasscode = 12345
            VMtoEmail = 0
            VMEmail = ""
            if (Ext.voicemail_toEmail == True):
                VMEmail = Ext.voicemail_email
                VMtoEmail = 1
            ExtActive = 1

            writer.writerow({
                'Extension Number': number,
                'Extension Name': name,
                'CIDName': cidName,
                'CIDNum': cidNumber,
                'E911CIDNum': E911Number,
                'DirectorySearchable': directorySearch,
                'DefaultNPA': defaultNPA,
                'Wrap Time': wrapTime,
                'Record Call': recordCall,
                'FwdTime': forwardTime,
                'Record Notify Email': recordNotify,
                'Roaming Passcode': roamingPass,
                'VM Extension Number': VMExtNumber,
                'VM Forward Email': VMEmail,
                'VM to Email Only': VMtoEmail,
                'VM Passcode': VMPasscode,
                'SIP Username': "",
                'SIP Password': "",
                'Extension Active': ExtActive
                })
    
    try:
        tempDel = Reports.objects.filter(company=tempComp, type='Extensions')
        for each in tempDel:
            each.delete()
            
    except Reports.DoesNotExist:
        print()
    
    x = Reports.objects.create(company=tempComp, document=toForm, type="Extensions")
    x.save()

    return HttpResponse("Done Report" + str(company))


@login_required
def CreatePortingForm(request):

    if request.session.get("company"):
        company = request.session.get("company")
        print(company)
    else:
        raise Http404

    template = os.path.join(settings.MEDIA_ROOT, "Reports", "Template_Porting.xlsx")

    tempComp = Company.objects.get(id=company)
    save_path = os.path.join(settings.MEDIA_ROOT, "Reports", str(tempComp.order) + "_Porting.xlsx")
    toForm = os.path.join("Reports", str(tempComp.order) + "_Porting.xlsx")

    copyfile(template, save_path)

    wb = load_workbook(save_path)
    ws = wb.active

    #First Part
    ws['E9'] = tempComp.company_name
    tempAddress = Address.objects.get(id=tempComp.site_address_id)
    if tempAddress.Suite == "":
        suiteHandler = ""
    else:
        suiteHandler = str(tempAddress.Suite) + " - "
    ws['E10'] = suiteHandler + tempAddress.StreetAddress
    ws['E11'] = tempComp.type
    ws['E13'] = tempComp.currentProvider


    ws['E16'] = tempComp.listing_name
    ws['E17'] = tempComp.type
    ws['E18'] = tempComp.category_listing
    ws['E19'] = tempComp.listing_phone

    tempAddress = Address.objects.get(id=tempComp.listing_address_id)
    res = tempAddress.StreetAddress.split(", ");
    if tempAddress.Suite == "":
        suiteHandler = ""
    else:
        suiteHandler = str(tempAddress.Suite) + " - "
        
    ws['E20'] = suiteHandler + res[0]
    ws['E23'] = res[1]
    ws['E24'] = res[2]
    ws['E25'] = tempAddress.Postal

    ws['A29'] = "Phone number(s) to be ported (If more than the below fields, please attach EXCEL SHEET with all the TNs)"
    counter = 30

    tempPort = Numbers.objects.filter(Type=1, company_id=tempComp)
    for each in tempPort:
        ws['A'+ str(counter)] = each.number
        counter = counter + 1
    
    counter = counter + 1
    ws['A'+ str(counter)] = "Phone number(s) to be disconnect (if applicable) (If more than the below fields, please attach EXCEL SHEET with all the TNs)"
    counter = counter + 1

    tempPort = Numbers.objects.filter(Type=0, company_id=tempComp)
    for each in tempPort:
        ws['A'+ str(counter)] = each.number
        counter = counter + 1
    
    
    counter = counter + 1
    ws['A'+ str(counter)] = "Authorized Customer Signiture"
    counter = counter + 1

    tempUpload = Uploads.objects.get(company_id=tempComp, type='signiture')
    img = openpyxl.drawing.image.Image(tempUpload.document)
    
    img.anchor = 'A' + str(counter)
    ws.add_image(img)
    
    counter = counter + 8
    #Printed Name
    printed_name = request.POST['input_name']
    ws['A'+ str(counter)] = "Authorized Printed Name (as per above signiture)"
    counter = counter + 1
    ws['A'+ str(counter)] = printed_name.upper()

    counter = counter + 2
    ws['A'+ str(counter)] = "Date"
    counter = counter + 1

    today = date.today()
    d1 = today.strftime("%d/%m/%Y")
    ws['A'+ str(counter)] = d1

    try:
        tempDel = Reports.objects.filter(company=tempComp, type='PortingForm')
        for each in tempDel:
            each.delete()
            
    except Reports.DoesNotExist:
        print()
    
    x = Reports.objects.create(company=tempComp, document=toForm, type="PortingForm")
    x.save()
    wb.save(save_path)
    print("Porting")
    return HttpResponse("Done")

@login_required
def CreatePBXForm(request):

    if request.session.get("company"):
            company = request.session.get("company")
            print(company)
    else:
        raise Http404

    template = os.path.join(settings.MEDIA_ROOT, "Reports", "Template_PBX.xlsx")

    tempComp = Company.objects.get(id=company)
    save_path = os.path.join(settings.MEDIA_ROOT, "Reports", str(tempComp.order) + "_PBX.xlsx")
    toForm = os.path.join("Reports", str(tempComp.order) + "_PBX.xlsx")

    copyfile(template, save_path)

    wb = load_workbook(save_path)
    ws = wb.active

    ws['B3'] = tempComp.company_name
    tempAddress = Address.objects.get(id=tempComp.site_address_id)
    res = tempAddress.StreetAddress.split(", ");

    if tempAddress.Suite == "":
        suiteHandler = ""
    else:
        suiteHandler = str(tempAddress.Suite) + " - "

    ws['B4'] = suiteHandler + res[0]
    ws['B5'] = res[1]
    ws['B6'] = res[2]
    ws['B7'] = tempAddress.Postal
    ws['B8'] = tempComp.listing_phone

    #Contact Name
    ws['B10'] = str(request.user.first_name) + " " + str(request.user.last_name)
    #Contact Email
    ws['B11'] = request.user.email
    #Contact Phone
    ws['B12'] = request.user.phone_number


    tempNumbers = Numbers.objects.filter(Type=1, company_id=tempComp)
    counter = 15
    for each in tempNumbers:
        
        ws['A' + str(counter + 1)] = "DID #"
        ws['A' + str(counter + 2)] = "Address"
        ws['A' + str(counter + 3)] = "City"
        ws['A' + str(counter + 4)] = "Province"
        ws['A' + str(counter + 5)] = "Postal Code"

        tempAddress = Address.objects.get(id=each.Address_911_id)
        res = tempAddress.StreetAddress.split(", ");

        if tempAddress.Suite == "":
            suiteHandler = ""
        else:
            suiteHandler = str(tempAddress.Suite) + " - "

        ws['B' + str(counter + 1)] = each.number
        ws['B' + str(counter + 2)] = suiteHandler + res[0]
        ws['B' + str(counter + 3)] = res[1]
        ws['B' + str(counter + 4)] = res[2]
        ws['B' + str(counter + 5)] = tempAddress.Postal

        counter = counter + 6
    
    try:
        tempDel = Reports.objects.filter(company=tempComp, type='PBX')
        for each in tempDel:
            each.delete()
           
    except Reports.DoesNotExist:
        print()
    

    x = Reports.objects.create(company=tempComp, document=toForm, type="PBX")
    x.save()
    wb.save(save_path)

    print("PBX")
    return HttpResponse("Done")
